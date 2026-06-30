"""
Extractor de XLSX (Fase 2).

Superficies de ocultamiento propias de las hojas de calculo:
  - Hojas ocultas o 'veryHidden'                 -> HIDDEN_SHEET
  - Filas / columnas ocultas                      -> HIDDEN_ROWCOL
  - Fuente del mismo color que el fondo (blanca)  -> NEAR_WHITE
  - Formato de numero ';;;' (celda visualmente en blanco) -> HIDDEN_FORMAT
  - Comentarios / notas de celda                  -> COMMENT
  - Nombres definidos                             -> DEFINED_NAME
  - Metadatos del libro                           -> METADATA

Usamos openpyxl: a diferencia de python-docx en Word, openpyxl SI expone de forma
fiable el estado de las hojas, el color de fuente y el formato de numero, asi que
no necesitamos bajar al XML crudo aqui.
"""

from __future__ import annotations

import openpyxl

from veilscan.core.models import Finding, HideReason, Severity, Technique, TextSpan
from veilscan.extractors.base import BaseExtractor, ExtractionResult

NEAR_WHITE_THRESHOLD = 230


def _argb_is_near_white(rgb: str | None) -> bool:
    """openpyxl entrega color como ARGB ('FFFFFFFF'). Tomamos los ultimos 6 digitos."""
    if not rgb or not isinstance(rgb, str):
        return False
    hexpart = rgb[-6:]
    if len(hexpart) != 6:
        return False
    try:
        r, g, b = int(hexpart[0:2], 16), int(hexpart[2:4], 16), int(hexpart[4:6], 16)
    except ValueError:
        return False
    return r >= NEAR_WHITE_THRESHOLD and g >= NEAR_WHITE_THRESHOLD and b >= NEAR_WHITE_THRESHOLD


class XlsxExtractor(BaseExtractor):
    extensions = (".xlsx", ".xlsm")

    def extract(self, path: str) -> ExtractionResult:
        result = ExtractionResult()
        wb = openpyxl.load_workbook(path, data_only=False)

        self._metadata(wb, result)
        self._defined_names(wb, result)

        for ws in wb.worksheets:
            sheet_hidden = ws.sheet_state in ("hidden", "veryHidden")
            if sheet_hidden:
                result.structural_findings.append(
                    Finding(
                        technique=Technique.HIDDEN_TEXT,
                        severity=Severity.MEDIUM,
                        title=f"Hoja oculta: '{ws.title}' (estado: {ws.sheet_state})",
                        location=f"sheet:{ws.title}",
                        evidence=f"sheet_state={ws.sheet_state}",
                        detail="La hoja no se muestra al abrir el libro pero su contenido si se extrae.",
                        hidden=True,
                    )
                )

            hidden_rows = {i for i, d in ws.row_dimensions.items() if d.hidden}
            hidden_cols = {c for c, d in ws.column_dimensions.items() if d.hidden}

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    text = str(cell.value).strip()
                    if not text:
                        continue

                    reason = None
                    if sheet_hidden:
                        reason = HideReason.HIDDEN_SHEET
                    elif cell.row in hidden_rows or cell.column_letter in hidden_cols:
                        reason = HideReason.HIDDEN_ROWCOL
                    elif cell.number_format and cell.number_format.replace(" ", "") == ";;;":
                        reason = HideReason.HIDDEN_FORMAT
                    elif cell.font and cell.font.color and _argb_is_near_white(
                        getattr(cell.font.color, "rgb", None)
                    ):
                        reason = HideReason.NEAR_WHITE

                    result.spans.append(
                        TextSpan(
                            text=text,
                            visible=reason is None,
                            location=f"{ws.title}!{cell.coordinate}",
                            reason=reason,
                        )
                    )

            # comentarios de celda
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment and cell.comment.text and cell.comment.text.strip():
                        result.spans.append(
                            TextSpan(
                                text=cell.comment.text.strip(),
                                visible=False,
                                location=f"{ws.title}!{cell.coordinate} (comment)",
                                reason=HideReason.COMMENT,
                            )
                        )

        wb.close()
        return result

    def _metadata(self, wb, result: ExtractionResult) -> None:
        props = wb.properties
        meta = {}
        for attr in ("title", "subject", "creator", "keywords", "description", "lastModifiedBy", "category"):
            val = getattr(props, attr, None)
            if val and str(val).strip():
                meta[attr] = str(val)
                result.spans.append(
                    TextSpan(text=str(val), visible=False,
                             location=f"metadata:{attr}", reason=HideReason.METADATA)
                )
        result.metadata = meta

    def _defined_names(self, wb, result: ExtractionResult) -> None:
        try:
            names = wb.defined_names
            items = names.values() if hasattr(names, "values") else names.definedName
        except Exception:
            return
        for dn in items:
            val = getattr(dn, "value", "") or getattr(dn, "attr_text", "")
            name = getattr(dn, "name", "?")
            if val and str(val).strip():
                result.spans.append(
                    TextSpan(text=str(val), visible=False,
                             location=f"defined_name:{name}", reason=HideReason.DEFINED_NAME)
                )
